#!/usr/bin/env python3
"""販売用 Excel データ集を作る。`npm run build` 後に実行する（out/csv/*/all.csv を読むので、
サイトのCSVと中身が必ず一致する）。出力はリポジトリに含めない（有料商品）。
  python3 scripts/build_xlsx.py  →  dist/iwate-data_YYYY-MM-DD.xlsx
"""
import csv, io, json, os, re, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.join(HERE, '..')
OUT_DIR = os.path.join(ROOT, 'out', 'csv')
DIST = os.path.join(ROOT, 'dist')
os.makedirs(DIST, exist_ok=True)
today = datetime.date.today().isoformat()
ds = json.load(open(os.path.join(ROOT, 'data', 'dataset.json'), encoding='utf-8'))
SOURCES = ds['sources']

# lib/csv.ts のレジストリからシート名と出典キーを拾う（並び順もそのまま）
ts = io.open(os.path.join(ROOT, 'lib', 'csv.ts'), encoding='utf-8').read()
FAM = re.findall(r"^  (\w+): \{ label: '([^']+)'", ts, re.M)
SRC_KEY = {'dental': 'dental', 'population': 'population', 'aging': 'census', 'work': 'census', 'building': 'building',
           'vital': 'vital', 'household': 'household', 'medical': 'medical', 'welfare': 'welfare', 'garbage': 'env',
           'economy': 'economy', 'school': 'school', 'jobless': 'jobless', 'education': 'education', 'farm': 'farm'}

FONT = 'Meiryo'
head_font = Font(name=FONT, bold=True, color='FFFFFF', size=10)
head_fill = PatternFill('solid', fgColor='0017C1')
body_font = Font(name=FONT, size=10)
thin = Side(style='thin', color='D8D8DB')

wb = Workbook()
ws = wb.active; ws.title = 'はじめに'
intro = [
    ['岩手県33市町村 統計データ集', ''],
    ['版', today],
    ['作成', 'いわてデータ（ビークプロモーション株式会社、盛岡市）'],
    ['URL', 'https://iwate-data.jp/data/'],
    ['', ''],
    ['内容', '政府統計（e-Stat）の公表値を岩手県33市町村×年で整理したもの。1分野＝1シート、縦持ち（ロング形式）。'],
    ['検算', '各分野で、33市町村の合計が総務省統計局などが公表する岩手県の値と一致することを機械的に確認しています（率・原単位を除く）。'],
    ['空欄', '秘匿（X）・非公表（...）・該当なし（-）・未調査の年は空欄。推計・按分・補完はしていません。'],
    ['合併', '旧滝沢村（03305）→滝沢市、旧藤沢町（03422）→一関市、旧川井村（03487）→宮古市 に合算しています。'],
    ['利用条件', 'CC BY 4.0。出典として「いわてデータ」を明記すれば商用を含め自由に利用できます。原データの著作権は各統計の作成機関に帰属します。'],
    ['注意', '「従業者数」（経済センサス：その市町村の事業所で働く人）と「就業者数」（国勢調査：その市町村に住む働く人）は別の概念です。'],
    ['', ''],
    ['シート一覧', ''],
]
for k, label in FAM:
    intro.append([label, f"出典: {SOURCES[SRC_KEY[k]]['name']}"])
for r in intro: ws.append(r)
ws['A1'].font = Font(name=FONT, bold=True, size=14)
for row in ws.iter_rows(min_row=2):
    for c in row: c.font = body_font; c.alignment = Alignment(wrap_text=True, vertical='top')
    row[0].font = Font(name=FONT, bold=True, size=10)
ws.column_dimensions['A'].width = 22; ws.column_dimensions['B'].width = 110

total_rows = 0
for k, label in FAM:
    path = os.path.join(OUT_DIR, k, 'all.csv')
    rows = list(csv.reader(io.open(path, encoding='utf-8-sig', newline='')))
    sh = wb.create_sheet(label[:31])
    for i, r in enumerate(rows):
        if i == 0:
            sh.append(r)
        else:
            out = []
            for j, v in enumerate(r):
                if j < 2: out.append(v)              # コード・市町村名は文字列のまま（先頭0を守る）
                elif v == '': out.append(None)
                else:
                    try: out.append(int(v)) if re.fullmatch(r'-?\d+', v) else out.append(float(v))
                    except ValueError: out.append(v)
            sh.append(out)
    total_rows += len(rows) - 1
    for c in sh[1]:
        c.font = head_font; c.fill = head_fill; c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sh.row_dimensions[1].height = 32
    for row in sh.iter_rows(min_row=2):
        for c in row:
            c.font = body_font; c.border = Border(bottom=thin)
            if isinstance(c.value, (int, float)) and c.column > 3: c.number_format = '#,##0.##' if isinstance(c.value, float) else '#,##0'
    sh.freeze_panes = 'D2'
    sh.auto_filter.ref = sh.dimensions
    widths = [10, 12, 8] + [max(12, min(28, len(h) * 1.6)) for h in rows[0][3:]]
    for j, w in enumerate(widths, 1): sh.column_dimensions[get_column_letter(j)].width = w

# 出典・注記シート
so = wb.create_sheet('出典・注記')
so.append(['分野', '統計名・表名', 'URL', '注記'])
for k, label in FAM:
    s = SOURCES[SRC_KEY[k]]
    so.append([label, s['name'], s['url'], s.get('note', '')])
for c in so[1]: c.font = head_font; c.fill = head_fill
for row in so.iter_rows(min_row=2):
    for c in row: c.font = body_font; c.alignment = Alignment(wrap_text=True, vertical='top')
for col, w in zip('ABCD', (18, 50, 40, 90)): so.column_dimensions[col].width = w

out = os.path.join(DIST, f'iwate-data_{today}.xlsx')
wb.save(out)
print('ok', out, os.path.getsize(out), 'bytes', total_rows, 'rows', len(FAM), 'sheets')
